"""
Combine multiple xlsx files into a single workbook with per-file tabs.

Preserves formatting, styles, merged cells, and column widths from each
source sheet.
"""

# Standard Library
import os

# PIP3 modules
import openpyxl
import openpyxl.styles
import openpyxl.worksheet.worksheet


#============================================

def copy_sheet(src_sheet: openpyxl.worksheet.worksheet.Worksheet, new_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
	"""
	Copy data, formatting, styles, and merged cells from src_sheet to new_sheet.

	Args:
		src_sheet: The source worksheet to copy from.
		new_sheet: The destination worksheet to copy to.
	"""
	# Copy column widths
	for col_letter, col_dim in src_sheet.column_dimensions.items():
		new_sheet.column_dimensions[col_letter].width = col_dim.width

	# Copy row heights
	for row_num, row_dim in src_sheet.row_dimensions.items():
		new_sheet.row_dimensions[row_num].height = row_dim.height

	# Copy data and formatting cell by cell
	for row in src_sheet.iter_rows():
		for cell in row:
			new_cell = new_sheet[cell.coordinate]
			new_cell.value = cell.value

			# Copy font style
			if cell.font:
				new_cell.font = openpyxl.styles.Font(
					name=cell.font.name,
					bold=cell.font.bold,
					italic=cell.font.italic,
					underline=cell.font.underline,
					size=cell.font.size,
					color=cell.font.color
				)

			# Copy fill/background
			if cell.fill:
				new_cell.fill = openpyxl.styles.PatternFill(
					fill_type=cell.fill.fill_type,
					start_color=cell.fill.start_color.index,
					end_color=cell.fill.end_color.index
				)

			# Copy cell border
			if cell.border:
				new_cell.border = openpyxl.styles.Border(
					left=cell.border.left,
					right=cell.border.right,
					top=cell.border.top,
					bottom=cell.border.bottom
				)

			# Copy text alignment
			if cell.alignment:
				new_cell.alignment = openpyxl.styles.Alignment(
					horizontal=cell.alignment.horizontal,
					vertical=cell.alignment.vertical,
					wrap_text=cell.alignment.wrap_text
				)

			# Copy number format
			if cell.number_format:
				new_cell.number_format = cell.number_format

	# Copy merged cell ranges
	for merged_range in src_sheet.merged_cells.ranges:
		new_sheet.merge_cells(str(merged_range))


#============================================

def merge_workbooks(input_files: list, output_file: str) -> None:
	"""
	Combine multiple xlsx files into a single workbook with each file as a tab.

	Args:
		input_files: List of Excel file paths to merge.
		output_file: Output Excel file path.
	"""
	if not input_files:
		raise ValueError("No input files provided.")

	# Create a new workbook; remove the default blank sheet
	output_wb = openpyxl.Workbook()
	output_wb.remove(output_wb.active)

	for file in input_files:
		wb = openpyxl.load_workbook(file)
		for sheet_name in wb.sheetnames:
			src_sheet = wb[sheet_name]
			# Excel sheet name max length is 31 characters
			new_sheet_name = (os.path.splitext(os.path.basename(file))[0])[:31]
			new_sheet = output_wb.create_sheet(title=new_sheet_name)
			copy_sheet(src_sheet, new_sheet)

	output_wb.save(output_file)
	print(f"Merged {len(input_files)} files into '{output_file}'.")


