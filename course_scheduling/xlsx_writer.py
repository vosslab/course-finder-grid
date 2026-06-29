"""
openpyxl rendering for the schedule grid.

Consumes the pure grid model (grid_model) and writes the spreadsheet: places
classes, merges duration cells, writes day/time labels, colors by course
identity, shades the common hour, and saves the workbook.
"""

# Standard Library
import datetime

# PIP3 modules
import openpyxl
import openpyxl.styles

# local repo modules
import course_scheduling.course_label
import course_scheduling.grid_model
import course_scheduling.schedule_colors


#============================================

def format_merged_cells(ws: openpyxl.worksheet.worksheet.Worksheet, waitlisted_labels: set | None = None, enrollment_ratios: dict | None = None) -> None:
	"""
	Format all merged cells: center align, vertical align, wrap text, and color with a black 2pt border.

	Args:
		ws (Worksheet): The worksheet object where formatting is applied.
	"""
	# Define the border style (2pt black border)
	black_border = openpyxl.styles.Border(
		left=openpyxl.styles.Side(border_style="medium", color="333333"),
		right=openpyxl.styles.Side(border_style="medium", color="333333"),
		top=openpyxl.styles.Side(border_style="medium", color="333333"),
		bottom=openpyxl.styles.Side(border_style="medium", color="333333")
	)

	# Default the optional inputs to empty containers when not provided
	if waitlisted_labels is None:
		waitlisted_labels = set()
	if enrollment_ratios is None:
		enrollment_ratios = {}

	# Loop through all merged cell ranges
	for merged_range in ws.merged_cells.ranges:
		# Get the top-left cell of the merged range
		top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
		course_label = top_left_cell.value
		course_label = course_label.replace('\n', ' ')
		if len(course_label) <= 1:
			continue
		label_key = course_scheduling.course_label.normalize_label_key(course_label)
		is_waitlisted = label_key in waitlisted_labels
		enrollment_ratio = enrollment_ratios.get(label_key)
		hex_color = course_scheduling.schedule_colors.label_to_pastel_hex(course_label, waitlisted=is_waitlisted, enrollment_ratio=enrollment_ratio)
		fill = openpyxl.styles.PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

		# Apply the formatting to the top-left cell
		top_left_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
		top_left_cell.fill = fill

		# Apply fill and border to every cell in the merged range, not just the top-left.
		for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row,
							min_col=merged_range.min_col, max_col=merged_range.max_col):
			for cell in row:
				cell.fill = fill  # Apply fill to all cells in the merged range
				cell.border = black_border  # Apply 2pt black border to all cells in the merged range


#============================================

def format_schedule_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, schedule_grid: dict, classes_list: list, columns_needed: dict, starting_columns: dict) -> None:
	"""
	Place each class into the grid and merge cells across its duration.

	Args:
		ws: The worksheet to format.
		schedule_grid: Dict of day to time-slot to class-label lists.
		classes_list: List of class dicts with Label, Days, Start, and End.
		columns_needed: Dict mapping day letter to number of columns.
		starting_columns: Dict mapping day letter to first column number.
	"""
	days_of_week = list(schedule_grid.keys())
	all_time_slots = list(schedule_grid[days_of_week[0]].keys())

	filled_grid_status = {}
	# Create an empty list for each day and each time slot
	total_columns_needed = sum(columns_needed.values())
	max_column = total_columns_needed + len(days_of_week) + 2
	for col in range(max_column):
		for row in range(len(all_time_slots)+2):
			filled_grid_status[(row,col)] = False

	for class_dict in classes_list:
		label = class_dict['Label']
		class_days = class_dict["Days"]
		begin_time = class_dict['Start']
		begin_row_number = course_scheduling.grid_model.time_slot_to_row_number(begin_time)
		end_time = class_dict['End']
		end_time = datetime.datetime.combine(datetime.datetime.today(), end_time) - datetime.timedelta(minutes=3)
		end_time = end_time.time()  # Convert back to time object if needed
		end_row_number = course_scheduling.grid_model.time_slot_to_row_number(end_time)

		for day in class_days:
			start_column = starting_columns[day]
			column_num = course_scheduling.grid_model.find_available_column(filled_grid_status, start_column, begin_row_number, end_row_number)

			# Enter the class label in the first cell (begin_row_number)
			ws.cell(row=begin_row_number, column=column_num).value = label

			# Merge the cells vertically to represent the class duration
			ws.merge_cells(start_row=begin_row_number, start_column=column_num,
						end_row=end_row_number, end_column=column_num)
			course_scheduling.grid_model.fill_the_status_grid(filled_grid_status, begin_row_number, end_row_number, column_num)


#============================================

def write_time_slots_to_column(ws: openpyxl.worksheet.worksheet.Worksheet, all_time_slots: list, column_number: int) -> None:
	"""
	Write the time-slot labels down a single column.

	Args:
		ws: The worksheet to write into.
		all_time_slots: List of 'HH:MM' time-slot strings.
		column_number: Column index that receives the time labels.
	"""
	for time_slot in all_time_slots:
		row_num = course_scheduling.grid_model.time_slot_to_row_number(time_slot)
		ws.cell(row=row_num, column=column_number).value = time_slot


#============================================

def write_spreadsheet_labels(ws: openpyxl.worksheet.worksheet.Worksheet, days_of_week: list, columns_needed: dict, starting_columns: dict, all_time_slots: list) -> None:
	"""
	Write the time column, day headers, and per-day time labels.

	Args:
		ws: The worksheet to write into.
		days_of_week: List of day letters in display order.
		columns_needed: Dict mapping day letter to number of columns.
		starting_columns: Dict mapping day letter to first column number.
		all_time_slots: List of 'HH:MM' time-slot strings.
	"""
	for time_slot in all_time_slots:
		row_num = course_scheduling.grid_model.time_slot_to_row_number(time_slot)
		ws.cell(row=row_num, column=1).value = time_slot

	# Merge cells for each day based on the starting_columns and columns_needed
	for day in days_of_week:
		num_columns = columns_needed[day]-1
		start_column = starting_columns[day]
		time_column = start_column-1
		ws.cell(row=1, column=time_column).value = "Time"
		write_time_slots_to_column(ws, all_time_slots, time_column)
		ws.cell(row=1, column=start_column).value = day

		# Merge the cells for this day if there are multiple columns needed
		if num_columns > 1:
			ws.merge_cells(start_row=1, start_column=start_column,
						end_row=1, end_column=start_column + num_columns - 1)


#============================================

def shade_common_hour(ws: openpyxl.worksheet.worksheet.Worksheet, starting_columns: dict, columns_needed: dict) -> None:
	"""
	Apply a low-saturation background to T and R common hour cells in the grid.

	Shades the 15-min slots covering 12:15-13:15 (grid-rounded common hour).
	Only shades empty cells; cells with course data keep their existing color.

	Args:
		ws: The openpyxl worksheet.
		starting_columns: Dict mapping day letter to first column number.
		columns_needed: Dict mapping day letter to number of columns for that day.
	"""
	# Light beige/tan fill with low saturation
	common_hour_hex = course_scheduling.schedule_colors.hls_to_hex(40, 0.90, 0.15)
	common_hour_fill = openpyxl.styles.PatternFill(
		start_color=common_hour_hex, end_color=common_hour_hex, fill_type="solid"
	)
	# Common hour grid slots: 12:15, 12:30, 12:45, 13:00, 13:15
	common_hour_slots = ["12:15", "12:30", "12:45", "13:00", "13:15"]
	for day in ("T", "R"):
		if day not in starting_columns:
			continue
		start_col = starting_columns[day]
		num_cols = columns_needed[day]
		for time_str in common_hour_slots:
			row_num = course_scheduling.grid_model.time_slot_to_row_number(time_str)
			for col in range(start_col, start_col + num_cols):
				cell = ws.cell(row=row_num, column=col)
				# Only shade empty cells
				if cell.value is None or cell.value == "":
					cell.fill = common_hour_fill


#============================================

def create_spreadsheet(schedule_grid: dict, sorted_classes: list, output_file: str) -> None:
	"""
	Create the schedule spreadsheet and save to Excel.

	Args:
		schedule_grid: Dictionary with class schedule grid.
		sorted_classes: List of classes sorted by Label.
	"""
	days_of_week = list(schedule_grid.keys())  # Ensure order if needed
	columns_needed, starting_columns = course_scheduling.grid_model.compute_columns_info(schedule_grid)
	all_time_slots = list(schedule_grid[days_of_week[0]].keys())

	# Initialize workbook and worksheet
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Schedule"

	write_spreadsheet_labels(ws, days_of_week, columns_needed, starting_columns, all_time_slots)

	# Center-align the entire first row (header row)
	for cell in ws[1]:  # This selects all cells in row 1
		cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

	# Populate the worksheet with schedule data
	format_schedule_sheet(ws, schedule_grid, sorted_classes, columns_needed, starting_columns)

	waitlisted_labels = set()
	enrollment_ratios = {}
	for class_row in sorted_classes:
		if class_row["Waitlisted"]:
			waitlisted_labels.add(course_scheduling.course_label.normalize_label_key(class_row["Label"]))
		enrollment_ratio = class_row["Enrollment_Ratio"]
		if enrollment_ratio is None:
			continue
		label_key = course_scheduling.course_label.normalize_label_key(class_row["Label"])
		current_ratio = enrollment_ratios.get(label_key)
		if current_ratio is None or enrollment_ratio > current_ratio:
			enrollment_ratios[label_key] = enrollment_ratio

	format_merged_cells(ws, waitlisted_labels, enrollment_ratios)

	# Shade common hour slots on T/R columns
	shade_common_hour(ws, starting_columns, columns_needed)

	# Save to an Excel file
	wb.save(output_file)
	print(f"Schedule saved to:\n\t{output_file}")
